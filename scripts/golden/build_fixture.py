"""Extract a bounded, diverse multi-category product fixture from Spec_cate_gia.xlsx
into the real DMX tool-payload shape. Deterministic (no randomness).

Output: data/dataset/golden/product-fixture.json
"""
import openpyxl, json, os, re

SRC = "data/dataset/Spec_cate_gia.xlsx"
OUT = "data/dataset/golden/product-fixture.json"
SNAPSHOT = "dmx-golden-fixture-2026-07-18"

# categories to include (diverse >= 8), with how many to sample each
CATS = {
    "Máy lạnh": 22, "Tủ Lạnh": 18, "Máy giặt": 16, "Màn hình máy tính": 16,
    "Máy tính để bàn": 12, "Máy tính bảng": 14, "Máy nước nóng": 10,
    "Đồng hồ thông minh": 12, "Máy rửa chén": 8, "Tủ mát, tủ đông": 8,
}

def num(x):
    if x is None: return None
    m = re.findall(r"\d+", str(x).replace(",", "").replace(".", ""))
    return int("".join(m)) if m else None

def price_of(row, idx):
    km = num(row[idx["giá khuyến mãi"]]) if "giá khuyến mãi" in idx else None
    gg = num(row[idx["giá gốc"]]) if "giá gốc" in idx else None
    sale = km or gg
    return gg, sale

def build():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    out = {"snapshot": SNAPSHOT, "source": "Spec_cate_gia.xlsx", "categories": {}}
    for cat, n in CATS.items():
        if cat not in wb.sheetnames:
            print("MISSING SHEET:", cat); continue
        ws = wb[cat]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
        idx = {h: i for i, h in enumerate(hdr)}
        # lowercase price header keys can vary in case across sheets
        for k in list(idx):
            if k.lower() in ("giá gốc", "giá khuyến mãi", "khuyến mãi quà"):
                idx[k.lower()] = idx[k]
        priced = []
        for r in rows[1:]:
            if not any(r): continue
            gg, sale = price_of(r, idx)
            if not sale or sale < 500000: continue
            priced.append((sale, gg, r))
        if not priced:
            print("NO PRICED ROWS:", cat); continue
        priced.sort(key=lambda t: t[0])
        # spread across price quantiles: pick n evenly across the sorted list
        picks = []
        L = len(priced)
        for k in range(n):
            picks.append(priced[min(L - 1, round(k * (L - 1) / max(1, n - 1)))])
        # dedup by row identity while preserving order
        seen = set(); uniq = []
        for sale, gg, r in picks:
            key = (r[idx.get("sku", 0)], sale)
            if key in seen: continue
            seen.add(key); uniq.append((sale, gg, r))
        products = []
        for i, (sale, gg, r) in enumerate(uniq):
            def g(col):
                j = idx.get(col)
                v = r[j] if j is not None and j < len(r) else None
                return None if v is None else str(v).strip()
            brand = g("brand") or "N/A"
            model = g("model_code") or g("sku") or f"{cat}-{i}"
            pid = num(g("productidweb")) or num(g("sku")) or (10_000_000 + i)
            # synthesize stock foils deterministically: index-based
            mod = i % 9
            stock = "Còn Hàng"
            if mod == 4: stock = "Hết Hàng"
            elif mod == 7: stock = "Liên hệ"   # unknown / contact
            # build spec string from all non-price, non-id columns that have values
            spec_bits = []
            for h in hdr:
                if h in ("model_code", "sku", "productidweb", "category_code",
                         "brand_id", "brand", "giá gốc", "giá khuyến mãi",
                         "khuyến mãi quà"):
                    continue
                v = g(h)
                if v and v.lower() not in ("none", "null", "không", ""):
                    spec_bits.append(f"{h}: {v}")
            promo = g("khuyến mãi quà")
            products.append({
                "product_id": pid,
                "productcode": g("sku"),
                "tên sản phẩm": f"{cat} {brand} {model}",
                "brand": brand,
                "category": cat,
                "Loại sản phẩm": cat + (", " + cat + " " + brand),
                "Trạng thái": stock,
                "Giá gốc": gg,
                "Giá khuyến mãi": sale,
                "Khuyến mãi": promo or "",
                "Thông tin sản phẩm": "; ".join(spec_bits),
                "raw_specs": {h: g(h) for h in hdr if g(h)},
            })
        out["categories"][cat] = products
        print(f"{cat}: {len(products)} products  price {products[0]['Giá khuyến mãi']:,}..{products[-1]['Giá khuyến mãi']:,}  stock foils: "
              f"{sum(1 for p in products if p['Trạng thái']!='Còn Hàng')}")
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    json.dump(out, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    total = sum(len(v) for v in out["categories"].values())
    print(f"\nWROTE {OUT}  categories={len(out['categories'])} products={total}")

    # Emit the FULL set of real product ids (productidweb) across every sheet, so
    # the eval harness can validate ids presented by the live agent (which loads
    # the whole xlsx, not just this sampled fixture).
    all_ids = set()
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        hdr = [str(c).strip() if c is not None else "" for c in next(rows, [])]
        if "productidweb" not in hdr:
            continue
        j = hdr.index("productidweb")
        for r in rows:
            if j < len(r):
                pid = num(r[j])
                if pid:
                    all_ids.add(pid)
    ALL = "data/dataset/golden/all-product-ids.json"
    json.dump(sorted(all_ids), open(ALL, "w", encoding="utf-8"))
    print(f"WROTE {ALL}  real product ids={len(all_ids)}")

if __name__ == "__main__":
    build()
