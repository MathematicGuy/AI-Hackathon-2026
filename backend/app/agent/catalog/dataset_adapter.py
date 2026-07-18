"""Excel dataset adapter — swap-ready for the future product database.

Loads `data/dataset/Spec_cate_gia.xlsx` (14 category sheets) into
`GenericProduct` records whose shape mirrors the committed logical format and
the future `products` table: mirror keys + an `attributes` dict preserving the
ORIGINAL Vietnamese column names and values + parsed price/promotion fields.
Replacing this adapter with a database-backed one must not change any consumer.
"""

import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

from backend.app.agent.catalog.promotions import PromotionInfo, extract_promotion

DEFAULT_DATASET_PATH = Path("data") / "dataset" / "Spec_cate_gia.xlsx"

MIRROR_KEYS = (
    "model_code",
    "sku",
    "productidweb",
    "category_code",
    "brand_id",
    "brand",
)


@dataclass(frozen=True, slots=True)
class GenericProduct:
    productidweb: str
    category_code: str
    category_name: str
    brand: str | None
    brand_id: str | None
    model_code: str | None
    sku: str | None
    attributes: dict[str, Any] = field(repr=False)
    promotion: PromotionInfo = field(repr=False)

    @property
    def list_price(self) -> int | None:
        return self.promotion.list_price

    @property
    def sale_price(self) -> int | None:
        return self.promotion.sale_price

    @property
    def effective_price(self) -> int | None:
        return self.promotion.effective_price

    @property
    def gift_promotion(self) -> str | None:
        return self.promotion.gift

    @property
    def name(self) -> str:
        brand = self.brand or ""
        model = self.model_code or self.sku or self.productidweb
        return f"{self.category_name} {brand} {model}".strip()


def _as_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


class ExcelDatasetAdapter:
    def __init__(self, path: Path | str | None = None) -> None:
        configured = path or os.environ.get("AGENT_DATASET_PATH") or DEFAULT_DATASET_PATH
        self.path = Path(configured)
        self._cache: list[GenericProduct] | None = None

    def load(self) -> list[GenericProduct]:
        if self._cache is not None:
            return self._cache
        if not self.path.exists():
            raise FileNotFoundError(f"dataset workbook not found: {self.path}")

        workbook = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
        products: list[GenericProduct] = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            header = [_as_text(cell) for cell in next(rows, ())]
            if not header or "productidweb" not in header:
                continue
            for row in rows:
                record = {
                    key: value
                    for key, value in zip(header, row)
                    if key is not None
                }
                productidweb = _as_text(record.get("productidweb"))
                if productidweb is None:
                    continue
                products.append(
                    GenericProduct(
                        productidweb=productidweb,
                        category_code=_as_text(record.get("category_code")) or "",
                        category_name=sheet_name,
                        brand=_as_text(record.get("brand")),
                        brand_id=_as_text(record.get("brand_id")),
                        model_code=_as_text(record.get("model_code")),
                        sku=_as_text(record.get("sku")),
                        attributes=record,
                        promotion=extract_promotion(record),
                    )
                )
        workbook.close()
        self._cache = products
        return products
