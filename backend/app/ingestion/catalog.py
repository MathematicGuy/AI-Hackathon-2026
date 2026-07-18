"""Product catalog ingestion.

Reads ``Spec_cate_gia.xlsx`` (14 category sheets) and upserts brands,
categories, and products. Every original column is preserved verbatim inside
``products.attributes`` (keys = exact original headers). Typed columns are
copies used only for indexing. Upserts are idempotent via ``content_hash``.
"""

from __future__ import annotations

import datetime as _dt
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook
from psycopg.types.json import Jsonb

from backend.app.ingestion.hashing import content_hash

CATALOG_FILENAME = "Spec_cate_gia.xlsx"

# Columns copied out of the original row into typed, indexable columns.
_MIRROR_KEYS = (
    "productidweb",
    "sku",
    "model_code",
    "category_code",
    "brand",
    "brand_id",
)


@dataclass
class CatalogStats:
    inserted: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0
    per_sheet: dict[str, dict[str, int]] = field(default_factory=dict)
    error_samples: list[str] = field(default_factory=list)

    @property
    def rows_ok(self) -> int:
        return self.inserted + self.updated

    def _bucket(self, sheet: str) -> dict[str, int]:
        return self.per_sheet.setdefault(
            sheet, {"inserted": 0, "updated": 0, "skipped": 0, "errors": 0}
        )

    def record(self, sheet: str, kind: str) -> None:
        setattr(self, kind, getattr(self, kind) + 1)
        self._bucket(sheet)[kind] += 1


def _jsonable(value: Any) -> Any:
    """Keep values verbatim but JSON-serializable."""
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped != "" else None
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        return value.isoformat()
    return value


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text != "" else None


def iter_rows(xlsx_path: Path) -> Iterator[tuple[str, dict[str, Any]]]:
    """Yield ``(sheet_name, record)`` where record keys are original headers."""
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            sheet = worksheet.title
            rows = worksheet.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue
            headers = [str(h).strip() if h is not None else None for h in header]
            for raw in rows:
                if raw is None:
                    continue
                if all(
                    cell is None or (isinstance(cell, str) and cell.strip() == "")
                    for cell in raw
                ):
                    continue
                record: dict[str, Any] = {}
                for head, cell in zip(headers, raw):
                    if head is None:
                        continue
                    record[head] = _jsonable(cell)
                yield sheet, record
    finally:
        workbook.close()


def _upsert_references(conn, rows: list[tuple[str, dict[str, Any]]]) -> None:
    brands: dict[str, str | None] = {}
    categories: dict[str, str] = {}
    for sheet, record in rows:
        brand = _text(record.get("brand"))
        if brand is not None and brand not in brands:
            brands[brand] = _text(record.get("brand_id"))
        category_code = _text(record.get("category_code"))
        if category_code is not None and category_code not in categories:
            categories[category_code] = sheet

    if brands:
        conn.cursor().executemany(
            """
            INSERT INTO brands (brand, brand_id) VALUES (%s, %s)
            ON CONFLICT (brand)
            DO UPDATE SET brand_id = COALESCE(brands.brand_id, EXCLUDED.brand_id)
            """,
            list(brands.items()),
        )
    if categories:
        conn.cursor().executemany(
            """
            INSERT INTO categories (category_code, sheet_name) VALUES (%s, %s)
            ON CONFLICT (category_code) DO UPDATE SET sheet_name = EXCLUDED.sheet_name
            """,
            list(categories.items()),
        )


def _existing_hashes(conn) -> dict[str, str]:
    rows = conn.execute("SELECT sku, content_hash FROM products").fetchall()
    return {row[0]: row[1] for row in rows}


def ingest_catalog(
    conn,
    xlsx_path: Path,
    *,
    source_file: str,
    import_run_id: int,
) -> CatalogStats:
    stats = CatalogStats()
    rows = list(iter_rows(xlsx_path))

    _upsert_references(conn, rows)
    conn.commit()

    existing = _existing_hashes(conn)
    seen: set[str] = set()
    to_insert: list[dict[str, Any]] = []
    to_update: list[dict[str, Any]] = []

    for sheet, record in rows:
        sku = _text(record.get("sku"))
        if sku is None:
            stats.record(sheet, "errors")
            if len(stats.error_samples) < 20:
                stats.error_samples.append(f"{sheet}: missing sku")
            continue
        if sku in seen:
            stats.record(sheet, "errors")
            if len(stats.error_samples) < 20:
                stats.error_samples.append(f"{sheet}: duplicate sku {sku}")
            continue
        seen.add(sku)

        digest = content_hash(record)
        params = {
            "sku": sku,
            "productidweb": _text(record.get("productidweb")),
            "model_code": _text(record.get("model_code")),
            "category_code": _text(record.get("category_code")),
            "brand": _text(record.get("brand")),
            "brand_id": _text(record.get("brand_id")),
            "sheet_name": sheet,
            "attributes": Jsonb(record),
            "source_file": source_file,
            "content_hash": digest,
            "import_run_id": import_run_id,
        }

        if sku not in existing:
            to_insert.append(params)
            stats.record(sheet, "inserted")
        elif existing[sku] != digest:
            to_update.append(params)
            stats.record(sheet, "updated")
        else:
            stats.record(sheet, "skipped")

    if to_insert:
        conn.cursor().executemany(
            """
            INSERT INTO products (
                sku, productidweb, model_code, category_code, brand, brand_id,
                sheet_name, attributes, source_file, content_hash, import_run_id
            ) VALUES (
                %(sku)s, %(productidweb)s, %(model_code)s, %(category_code)s,
                %(brand)s, %(brand_id)s, %(sheet_name)s, %(attributes)s,
                %(source_file)s, %(content_hash)s, %(import_run_id)s
            )
            """,
            to_insert,
        )
    if to_update:
        conn.cursor().executemany(
            """
            UPDATE products SET
                productidweb = %(productidweb)s,
                model_code = %(model_code)s,
                category_code = %(category_code)s,
                brand = %(brand)s,
                brand_id = %(brand_id)s,
                sheet_name = %(sheet_name)s,
                attributes = %(attributes)s,
                source_file = %(source_file)s,
                content_hash = %(content_hash)s,
                import_run_id = %(import_run_id)s,
                updated_at = now()
            WHERE sku = %(sku)s
            """,
            to_update,
        )
    conn.commit()
    return stats
